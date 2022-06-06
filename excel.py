import openpyxl
from openpyxl.styles import PatternFill

def write():
    #新規作成
    wb = openpyxl.Workbook()

    # 既存ファイルオープン
    #wb = openpyxl.load_workbook(filename="sample.xlsx")

    # book内のsheet名称一覧を取得する
    sheet_names = wb.sheetnames

    # sheet名称一覧 表示
    print (sheet_names) #debug

    # シート設定
    sheet = wb.worksheets[0]

    # シートのタイトル変更
    sheet.title = "sample1"

    # シート追加
    new_sheet = wb.create_sheet("sample2")

    # book内のsheet名称一覧を取得する
    sheet_names = wb.sheetnames

    # sheet名称一覧 表示
    print (sheet_names) #debug

    # セルA1への書き込み
    sheet.cell(row=1, column=1).value = 12

    # セルA2への書き込み
    sheet['A2'] = 34

    # セルA3にセルA1の値を書き込み
    sheet.cell(row=3, column=1).value = '=A1'

    # セル列を取得(ここではA)
    columns_str = openpyxl.utils.get_column_letter(1)

    # セル番号を追加(上からAを取得して2を加えるためA2)
    cell_str = columns_str + '2'

    # セルA4の値はセルA2(cell_str)1と同じ値
    sheet.cell(row=4, column=1).value = '=' + cell_str

    # Excelの関数も使用可能
    sheet.cell(row=5, column=1).value = '=SUM(A1:A4)'

    # オートフィルタ
    sheet.auto_filter.ref = 'A1:H1'

    # 1行目をウィンドウ固定
    sheet.freeze_panes = 'A2'

    # 色(前景色)を付ける
    fill = PatternFill(patternType='solid', fgColor='d3d3d3') #灰色
    sheet.cell(row=1, column=2).fill = fill
    fill = PatternFill(patternType='solid', fgColor='ffffcc') #黄色
    sheet.cell(row=2, column=2).fill = fill
    fill = PatternFill(patternType='solid', fgColor='ccffcc') #薄緑
    sheet.cell(row=3, column=2).fill = fill
    fill = PatternFill(patternType='solid', fgColor='33ffcc') #青緑
    sheet.cell(row=4, column=2).fill = fill
    fill = PatternFill(patternType='solid', fgColor='ffffdd') #薄黄
    sheet.cell(row=5, column=2).fill = fill
    fill = PatternFill(patternType='solid', fgColor='ddffff') #薄青緑
    sheet.cell(row=6, column=2).fill = fill
    fill = PatternFill(patternType='solid', fgColor='ffddff') #薄桃
    sheet.cell(row=7, column=2).fill = fill
    fill = PatternFill(patternType='solid', fgColor='00ee00') #緑
    sheet.cell(row=8, column=2).fill = fill
    fill = PatternFill(patternType='solid', fgColor='ffdbc9') #薄橙
    sheet.cell(row=9, column=2).fill = fill
    fill = PatternFill(patternType='solid', fgColor='ffd5ec') #薄赤
    sheet.cell(row=10, column=2).fill = fill
    fill = PatternFill(patternType='solid', fgColor='d9e5ff') #薄青
    sheet.cell(row=11, column=2).fill = fill
    # 色(背景色)を付ける
    fill = PatternFill(patternType='solid', fgColor='ffdead', bgColor='ffdead') #薄橙
    sheet.cell(row=12, column=2).fill = fill
    fill = PatternFill(patternType='solid', bgColor='d9e5ff') #bgColor指定すると変わらず黒になってしまう
    sheet.cell(row=13, column=2).fill = fill
    #の場合に色が変わらない...黒になったり，前景色と同じにならない

    # 保存
    wb.save("sample.xlsx")

    # 終了
    wb.close()

if __name__ == '__main__':
    write()
